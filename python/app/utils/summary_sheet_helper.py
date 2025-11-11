"""
Helper function to generate summary sheet Excel file matching Node.js implementation
This creates Excel files with proper formulas, formatting, and all sheets matching the Node.js version
"""
import os
from datetime import datetime
from sqlalchemy import create_engine, text as sync_text
from app.config.settings import settings
import logging
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Constants matching Node.js implementation
THREE_PO_VS_POS_REASON_LIST = [
    "NET_AMOUNT_ZOMATO_POS_MISMATCH",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH",
    "COMMISSION_VALUE_ZOMATO_POS_MISMATCH",
    "PG_CHARGE_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH",
    "PG_APPLIED_ON_ZOMATO_POS_MISMATCH",
    "TAXES_ZOMATO_FEE_ZOMATO_POS_MISMATCH",
    "TDS_AMOUNT_ZOMATO_POS_MISMATCH",
    "FINAL_AMOUNT_ZOMATO_POS_MISMATCH",
]

THREEPO_EXCEL_MISMATCH_REASONS = {
    "NET_AMOUNT_ZOMATO_POS_MISMATCH": "Net Amount mismatch",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH": "Tax Paid by Customer mismatch",
    "COMMISSION_VALUE_ZOMATO_POS_MISMATCH": "Commission Value mismatch",
    "PG_CHARGE_ZOMATO_POS_MISMATCH": "PG Charge mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH": "Net Amount & Tax Paid by Customer mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH": "Net Amount & Commission Value mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Net Amount & PG Charge mismatch",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH": "Tax Paid by Customer & Commission Value mismatch",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Tax Paid by Customer & PG Charge mismatch",
    "COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Commission Value & PG Charge mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH": "Net Amount, Tax Paid by Customer & Commission Value mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Net Amount, Tax Paid by Customer & PG Charge mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Net Amount, Commission Value & PG Charge mismatch",
    "TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Tax Paid by Customer, Commission Value & PG Charge mismatch",
    "NET_AMOUNT_ZOMATO_POS_MISMATCH,TAX_PAID_BY_CUSTOMER_ZOMATO_POS_MISMATCH,COMMISSION_VALUE_ZOMATO_POS_MISMATCH,PG_CHARGE_ZOMATO_POS_MISMATCH": "Net Amount, Tax Paid by Customer, Commission Value & PG Charge mismatch",
    "PG_APPLIED_ON_ZOMATO_POS_MISMATCH": "PG Applied On mismatch",
    "TAXES_ZOMATO_FEE_ZOMATO_POS_MISMATCH": "Taxes Zomato Fee mismatch",
    "TDS_AMOUNT_ZOMATO_POS_MISMATCH": "TDS Amount mismatch",
    "FINAL_AMOUNT_ZOMATO_POS_MISMATCH": "Final Amount mismatch",
}

# Column definitions matching Node.js
POS_VS_ZOMATO_COLUMNS = [
    "pos_order_id", "zomato_order_id", "store_name", "order_date",
    "pos_net_amount", "zomato_net_amount", "pos_vs_zomato_net_amount_delta",
    "pos_tax_paid_by_customer", "zomato_tax_paid_by_customer", "pos_vs_zomato_tax_paid_by_customer_delta",
    "pos_commission_value", "zomato_commission_value", "pos_vs_zomato_commission_value_delta",
    "pos_pg_applied_on", "zomato_pg_applied_on", "pos_vs_zomato_pg_applied_on_delta",
    "pos_pg_charge", "zomato_pg_charge", "pos_vs_zomato_pg_charge_delta",
    "pos_taxes_zomato_fee", "zomato_taxes_zomato_fee", "pos_vs_zomato_taxes_zomato_fee_delta",
    "pos_tds_amount", "zomato_tds_amount", "pos_vs_zomato_tds_amount_delta",
    "pos_final_amount", "zomato_final_amount", "pos_vs_zomato_final_amount_delta",
    "reconciled_status", "reconciled_amount", "unreconciled_amount",
    "pos_vs_zomato_reason", "order_status_pos"
]

ZOMATO_VS_POS_COLUMNS = [
    "zomato_order_id", "pos_order_id", "order_date", "store_name",
    "zomato_net_amount", "pos_net_amount", "zomato_vs_pos_net_amount_delta",
    "zomato_tax_paid_by_customer", "pos_tax_paid_by_customer", "zomato_vs_pos_tax_paid_by_customer_delta",
    "zomato_commission_value", "pos_commission_value", "zomato_vs_pos_commission_value_delta",
    "zomato_pg_applied_on", "pos_pg_applied_on", "zomato_vs_pos_pg_applied_on_delta",
    "zomato_pg_charge", "pos_pg_charge", "zomato_vs_pos_pg_charge_delta",
    "zomato_taxes_zomato_fee", "pos_taxes_zomato_fee", "zomato_vs_pos_taxes_zomato_fee_delta",
    "zomato_tds_amount", "pos_tds_amount", "zomato_vs_pos_tds_amount_delta",
    "zomato_final_amount", "pos_final_amount", "zomato_vs_pos_final_amount_delta",
    "calculated_zomato_net_amount", "calculated_zomato_tax_paid_by_customer",
    "calculated_zomato_commission_value", "calculated_zomato_pg_applied_on",
    "calculated_zomato_pg_charge", "calculated_zomato_taxes_zomato_fee",
    "calculated_zomato_tds_amount", "calculated_zomato_final_amount",
    "fixed_credit_note_amount", "fixed_pro_discount_passthrough",
    "fixed_customer_discount", "fixed_rejection_penalty_charge",
    "fixed_user_credits_charge", "fixed_promo_recovery_adj",
    "fixed_icecream_handling", "fixed_icecream_deductions",
    "fixed_order_support_cost", "fixed_merchant_delivery_charge",
    "reconciled_status", "reconciled_amount", "unreconciled_amount",
    "zomato_vs_pos_reason", "order_status_zomato"
]

ORDERS_NOT_IN_POS_COLUMNS = [
    "zomato_order_id", "order_date", "store_name",
    "zomato_net_amount", "zomato_tax_paid_by_customer", "zomato_commission_value",
    "zomato_pg_applied_on", "zomato_pg_charge", "zomato_taxes_zomato_fee",
    "zomato_tds_amount", "zomato_final_amount",
    "calculated_zomato_net_amount", "calculated_zomato_tax_paid_by_customer",
    "calculated_zomato_commission_value", "calculated_zomato_pg_applied_on",
    "calculated_zomato_pg_charge", "calculated_zomato_taxes_zomato_fee",
    "calculated_zomato_tds_amount", "calculated_zomato_final_amount",
    "fixed_credit_note_amount", "fixed_pro_discount_passthrough",
    "fixed_customer_discount", "fixed_rejection_penalty_charge",
    "fixed_user_credits_charge", "fixed_promo_recovery_adj",
    "fixed_icecream_handling", "fixed_icecream_deductions",
    "fixed_order_support_cost", "fixed_merchant_delivery_charge",
    "reconciled_status", "reconciled_amount", "unreconciled_amount",
    "zomato_vs_pos_reason", "order_status_zomato"
]

ORDERS_NOT_IN_3PO_COLUMNS = [
    "pos_order_id", "store_name", "order_date",
    "pos_net_amount", "pos_tax_paid_by_customer", "pos_commission_value",
    "pos_pg_applied_on", "pos_pg_charge", "pos_taxes_zomato_fee",
    "pos_tds_amount", "pos_final_amount",
    "reconciled_status", "reconciled_amount", "unreconciled_amount",
    "pos_vs_zomato_reason", "order_status_pos"
]

TEXT_COLUMNS_POS = ["pos_order_id", "zomato_order_id", "store_name", "order_date", "reconciled_status", "pos_vs_zomato_reason", "order_status_pos"]
TEXT_COLUMNS_ZOMATO = ["zomato_order_id", "pos_order_id", "store_name", "order_date", "reconciled_status", "zomato_vs_pos_reason", "order_status_zomato"]


def format_date(date_obj):
    """Format date like Node.js dayjs: MMM DD, YYYY"""
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.strptime(date_obj, "%Y-%m-%d").date()
        except:
            pass
    if hasattr(date_obj, 'strftime'):
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        return f"{months[date_obj.month - 1]} {date_obj.day:02d}, {date_obj.year}"
    return str(date_obj)


def apply_border(cell):
    """Apply thin border to cell"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def apply_outer_border(worksheet, start_row, end_row, start_col, end_col):
    """Apply outer border to range"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border


def generate_summary_sheet(workbook, start_date_dt, end_date_dt):
    """Generate Summary sheet matching Node.js generateSummarySheetForZomato"""
    worksheet = workbook.create_sheet("Summary", 0)  # Insert as first sheet
    
    # Row 1: Debtor Name
    worksheet['A1'] = "Debtor Name"
    worksheet['A1'].font = Font(bold=True)
    worksheet['B1'] = "Zomato"
    
    # Row 2: Recon Period
    worksheet['A2'] = "Recon Period"
    worksheet['A2'].font = Font(bold=True)
    start_date_str = format_date(start_date_dt)
    end_date_str = format_date(end_date_dt)
    worksheet['B2'] = f"{start_date_str} - {end_date_str}"
    
    # Row 3: Headers
    worksheet['B3'] = "No. of orders"
    worksheet['B3'].font = Font(bold=True)
    worksheet['C3'] = "POS Amount"
    worksheet['C3'].font = Font(bold=True)
    
    # Rows 4-6: POS Sale summary
    worksheet['A4'] = "POS Sale as per Business Date (S1+S2)"
    worksheet['A4'].font = Font(bold=True)
    worksheet['B4'] = f"=COUNTA('Zomato POS vs 3PO'!A2:A1048576)"
    worksheet['C4'] = f"=SUM('Zomato POS vs 3PO'!E2:E1048576)"
    
    worksheet['A5'] = "POS Sale as per Transaction Date (S1)"
    worksheet['A5'].font = Font(bold=True)
    worksheet['B5'] = f"=COUNTA('Zomato POS vs 3PO'!A2:A1048576)"
    worksheet['C5'] = f"=SUM('Zomato POS vs 3PO'!E2:E1048576)"
    
    worksheet['A6'] = "Difference in POS Sale that falls in subsequent time period (S2)"
    worksheet['A6'].font = Font(bold=True)
    worksheet['B6'] = "=B4-B5"
    worksheet['C6'] = "=C4-C5"
    
    # Empty row
    worksheet['A8'] = "POS Sale as per Business Date (S1+S2)"
    worksheet['A8'].font = Font(bold=True)
    apply_border(worksheet['A8'])
    
    # Merge cells for headers
    worksheet.merge_cells('B8:F8')
    worksheet['B8'] = "As per POS data (POS vs 3PO)"
    worksheet['B8'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B8'].font = Font(bold=True)
    apply_border(worksheet['B8'])
    
    worksheet.merge_cells('G8:K8')
    worksheet['G8'] = "As per 3PO Data (3PO vs POS)"
    worksheet['G8'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['G8'].font = Font(bold=True)
    apply_border(worksheet['G8'])
    
    # Row 9: Column headers
    headers_row9 = ["Parameters", "No. of orders", "POS Amount/Calculated", "3PO Amount/Actual", 
                   "Diff. in Amount", "Amount Receivable", "No. of orders", "3PO Amount/Actual",
                   "POS Amount/Calculated", "Diff. in Amount", "Amount Receivable"]
    for col_idx, header in enumerate(headers_row9, start=1):
        cell = worksheet.cell(row=9, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        apply_border(cell)
    
    apply_outer_border(worksheet, 9, 9, 2, 6)
    apply_outer_border(worksheet, 9, 9, 7, 11)
    
    # Row 10: DELIVERED
    row10_data = [
        "DELIVERED (As per transaction date)",
        f"=COUNTA('Zomato POS vs 3PO'!A2:A1048576)",
        f"=SUM('Zomato POS vs 3PO'!E2:E1048576)",
        f"=SUM('Zomato POS vs 3PO'!F2:F1048576)",
        "=C10-D10",
        f"=SUM('Zomato POS vs 3PO'!Z2:Z1048576)",
        f"=COUNTA('Zomato 3PO vs POS'!A2:A1048576)",
        f"=SUM('Zomato 3PO vs POS'!E2:E1048576)",
        f"=SUM('Zomato 3PO vs POS'!F2:F1048576)",
        "=H10-I10",
        f"=SUM('Zomato 3PO vs POS'!Z2:Z1048576)"
    ]
    for col_idx, value in enumerate(row10_data, start=1):
        cell = worksheet.cell(row=10, column=col_idx)
        if isinstance(value, str) and value.startswith('='):
            cell.value = value  # Formula
        else:
            cell.value = value
        cell.font = Font(bold=True)
    
    # Row 11: SALE
    worksheet['A11'] = "SALE"
    worksheet['A11'].alignment = Alignment(horizontal='right')
    worksheet['G11'] = f"=COUNTIFS('Zomato 3PO vs POS'!A2:A1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"sale\")"
    worksheet['H11'] = f"=SUMIFS('Zomato 3PO vs POS'!E2:E1048576,'Zomato 3PO vs POS'!E2:E1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"sale\")"
    worksheet['I11'] = f"=SUMIFS('Zomato 3PO vs POS'!F2:F1048576,'Zomato 3PO vs POS'!F2:F1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"sale\")"
    worksheet['J11'] = "=H11-I11"
    worksheet['K11'] = f"=SUMIFS('Zomato 3PO vs POS'!Z2:Z1048576,'Zomato 3PO vs POS'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"sale\")"
    
    # Row 12: ADDITION
    worksheet['A12'] = "ADDITION"
    worksheet['A12'].alignment = Alignment(horizontal='right')
    worksheet['G12'] = f"=COUNTIFS('Zomato 3PO vs POS'!A2:A1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"addition\")"
    worksheet['H12'] = f"=SUMIFS('Zomato 3PO vs POS'!E2:E1048576,'Zomato 3PO vs POS'!E2:E1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"addition\")"
    worksheet['I12'] = f"=SUMIFS('Zomato 3PO vs POS'!F2:F1048576,'Zomato 3PO vs POS'!F2:F1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"addition\")"
    worksheet['J12'] = "=H12-I12"
    worksheet['K12'] = f"=SUMIFS('Zomato 3PO vs POS'!Z2:Z1048576,'Zomato 3PO vs POS'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS'!AY2:AY1048576,\"addition\")"
    
    # Row 13: REFUND
    worksheet['A13'] = "REFUND"
    worksheet['A13'].alignment = Alignment(horizontal='right')
    worksheet['G13'] = f"=COUNTIFS('Zomato 3PO vs POS Refund'!A2:A1048576,\"<>\",'Zomato 3PO vs POS Refund'!AY2:AY1048576,\"refund\")"
    worksheet['H13'] = f"=SUMIFS('Zomato 3PO vs POS Refund'!E2:E1048576,'Zomato 3PO vs POS Refund'!E2:E1048576,\"<>\",'Zomato 3PO vs POS Refund'!AY2:AY1048576,\"refund\")"
    worksheet['I13'] = f"=SUMIFS('Zomato 3PO vs POS Refund'!F2:F1048576,'Zomato 3PO vs POS Refund'!F2:F1048576,\"<>\",'Zomato 3PO vs POS Refund'!AY2:AY1048576,\"refund\")"
    worksheet['J13'] = "=H13-I13"
    worksheet['K13'] = f"=SUMIFS('Zomato 3PO vs POS Refund'!Z2:Z1048576,'Zomato 3PO vs POS Refund'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS Refund'!AY2:AY1048576,\"refund\")"
    
    # Row 15: Reconciled Orders
    row15_data = [
        "Reconciled Orders",
        f"=COUNTIF('Zomato POS vs 3PO'!AC2:AC1048576,\"RECONCILED\")",
        f"=SUMIFS('Zomato POS vs 3PO'!E2:E1048576,'Zomato POS vs 3PO'!E2:E1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"RECONCILED\")",
        f"=SUMIFS('Zomato POS vs 3PO'!F2:F1048576,'Zomato POS vs 3PO'!F2:F1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"RECONCILED\")",
        "=C15-D15",
        f"=SUMIFS('Zomato POS vs 3PO'!Z2:Z1048576,'Zomato POS vs 3PO'!Z2:Z1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"RECONCILED\")",
        f"=COUNTIF('Zomato 3PO vs POS'!AU2:AU1048576,\"RECONCILED\")",
        f"=SUMIFS('Zomato 3PO vs POS'!E2:E1048576,'Zomato 3PO vs POS'!E2:E1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"RECONCILED\")",
        f"=SUMIFS('Zomato 3PO vs POS'!F2:F1048576,'Zomato 3PO vs POS'!F2:F1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"RECONCILED\")",
        "=H15-I15",
        f"=SUMIFS('Zomato 3PO vs POS'!Z2:Z1048576,'Zomato 3PO vs POS'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"RECONCILED\")"
    ]
    for col_idx, value in enumerate(row15_data, start=1):
        cell = worksheet.cell(row=15, column=col_idx)
        cell.value = value
        cell.font = Font(bold=True)
    
    # Row 16: Cancelled by Merchant and found in POS
    worksheet['A16'] = "Cancelled by Merchant and found in POS"
    worksheet['A16'].alignment = Alignment(horizontal='right')
    for col in range(2, 12):
        worksheet.cell(row=16, column=col).value = 0
    
    # Row 17: Cancelled by Merchant and not found in POS
    worksheet['A17'] = "Cancelled by Merchant and not found in POS"
    worksheet['A17'].alignment = Alignment(horizontal='right')
    for col in range(2, 12):
        worksheet.cell(row=17, column=col).value = 0
    
    # Row 18: Unreconciled Orders
    row18_data = [
        "Unreconciled Orders",
        f"=COUNTIF('Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\")",
        f"=SUMIFS('Zomato POS vs 3PO'!E2:E1048576,'Zomato POS vs 3PO'!E2:E1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\")",
        f"=SUMIFS('Zomato POS vs 3PO'!F2:F1048576,'Zomato POS vs 3PO'!F2:F1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\")",
        "=C18-D18",
        f"=SUMIFS('Zomato POS vs 3PO'!Z2:Z1048576,'Zomato POS vs 3PO'!Z2:Z1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\")",
        f"=COUNTIF('Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\")",
        f"=SUMIFS('Zomato 3PO vs POS'!E2:E1048576,'Zomato 3PO vs POS'!E2:E1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\")",
        f"=SUMIFS('Zomato 3PO vs POS'!F2:F1048576,'Zomato 3PO vs POS'!F2:F1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\")",
        "=H18-I18",
        f"=SUMIFS('Zomato 3PO vs POS'!Z2:Z1048576,'Zomato 3PO vs POS'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\")"
    ]
    for col_idx, value in enumerate(row18_data, start=1):
        cell = worksheet.cell(row=18, column=col_idx)
        cell.value = value
        cell.font = Font(bold=True)
    
    # Row 19: Order Not found in 3PO/POS
    row19_data = [
        "Order Not found in 3PO/POS",
        f"=COUNTIFS('Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!B2:B1048576,\"\")",
        f"=SUMIFS('Zomato POS vs 3PO'!E2:E1048576,'Zomato POS vs 3PO'!E2:E1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!B2:B1048576,\"\")",
        f"=SUMIFS('Zomato POS vs 3PO'!F2:F1048576,'Zomato POS vs 3PO'!F2:F1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!B2:B1048576,\"\")",
        "=C19-D19",
        f"=SUMIFS('Zomato POS vs 3PO'!Z2:Z1048576,'Zomato POS vs 3PO'!Z2:Z1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!B2:B1048576,\"\")",
        f"=COUNTIFS('Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!B2:B1048576,\"\")",
        f"=SUMIFS('Zomato 3PO vs POS'!E2:E1048576,'Zomato 3PO vs POS'!E2:E1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!B2:B1048576,\"\")",
        f"=SUMIFS('Zomato 3PO vs POS'!F2:F1048576,'Zomato 3PO vs POS'!F2:F1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!B2:B1048576,\"\")",
        "=H19-I19",
        f"=SUMIFS('Zomato 3PO vs POS'!Z2:Z1048576,'Zomato 3PO vs POS'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!B2:B1048576,\"\")"
    ]
    for col_idx, value in enumerate(row19_data, start=1):
        cell = worksheet.cell(row=19, column=col_idx)
        cell.value = value
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='right')
    
    # Add mismatch reasons (rows 20+)
    for index, reason in enumerate(THREE_PO_VS_POS_REASON_LIST):
        row_num = 20 + index
        reason_text = THREEPO_EXCEL_MISMATCH_REASONS.get(reason, reason)
        worksheet.cell(row=row_num, column=1).value = reason_text
        worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
        
        # POS vs 3PO formulas
        worksheet.cell(row=row_num, column=2).value = f"=COUNTIFS('Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!AF2:AF1048576,\"{reason}\")"
        worksheet.cell(row=row_num, column=3).value = f"=SUMIFS('Zomato POS vs 3PO'!E2:E1048576,'Zomato POS vs 3PO'!E2:E1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!AF2:AF1048576,\"{reason}\")"
        worksheet.cell(row=row_num, column=4).value = f"=SUMIFS('Zomato POS vs 3PO'!F2:F1048576,'Zomato POS vs 3PO'!F2:F1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!AF2:AF1048576,\"{reason}\")"
        worksheet.cell(row=row_num, column=5).value = f"=C{row_num}-D{row_num}"
        worksheet.cell(row=row_num, column=6).value = f"=SUMIFS('Zomato POS vs 3PO'!Z2:Z1048576,'Zomato POS vs 3PO'!Z2:Z1048576,\"<>\",'Zomato POS vs 3PO'!AC2:AC1048576,\"UNRECONCILED\",'Zomato POS vs 3PO'!AF2:AF1048576,\"{reason}\")"
        
        # 3PO vs POS formulas
        worksheet.cell(row=row_num, column=7).value = f"=COUNTIFS('Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!AX2:AX1048576,\"{reason}\")"
        worksheet.cell(row=row_num, column=8).value = f"=SUMIFS('Zomato 3PO vs POS'!E2:E1048576,'Zomato 3PO vs POS'!E2:E1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!AX2:AX1048576,\"{reason}\")"
        worksheet.cell(row=row_num, column=9).value = f"=SUMIFS('Zomato 3PO vs POS'!F2:F1048576,'Zomato 3PO vs POS'!F2:F1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!AX2:AX1048576,\"{reason}\")"
        worksheet.cell(row=row_num, column=10).value = f"=H{row_num}-I{row_num}"
        worksheet.cell(row=row_num, column=11).value = f"=SUMIFS('Zomato 3PO vs POS'!Z2:Z1048576,'Zomato 3PO vs POS'!Z2:Z1048576,\"<>\",'Zomato 3PO vs POS'!AU2:AU1048576,\"UNRECONCILED\",'Zomato 3PO vs POS'!AX2:AX1048576,\"{reason}\")"
    
    # Apply borders
    apply_outer_border(worksheet, 10, 38, 1, 1)
    apply_outer_border(worksheet, 10, 38, 2, 6)
    apply_outer_border(worksheet, 10, 38, 7, 11)


def create_data_sheet(workbook, sheet_name, columns, query_result, text_columns):
    """Create a data sheet with headers and data"""
    worksheet = workbook.create_sheet(sheet_name)
    
    # Add headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        apply_border(cell)
    
    # Add data rows
    for row_idx, record in enumerate(query_result, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = record.get(col_name) if isinstance(record, dict) else getattr(record, col_name, None)
            
            # Convert to number if not in text columns
            if col_name not in text_columns and value is not None:
                try:
                    if isinstance(value, str) and value.replace('.', '', 1).replace('-', '', 1).isdigit():
                        value = float(value)
                    elif not isinstance(value, (str, bool)):
                        value = float(value) if value else 0
                except (ValueError, TypeError):
                    pass
            
            cell.value = value
            apply_border(cell)
    
    # Auto-adjust column widths
    for col_idx in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in worksheet[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_summary_sheet_to_file(
    filepath: str,
    start_date_dt,
    end_date_dt,
    store_codes: list,
    progress_callback=None
):
    """
    Generate summary sheet Excel file matching Node.js implementation
    Creates Summary sheet with formulas and all data sheets
    """
    # Create sync engine for pandas operations
    sync_db_url = f"mysql+pymysql://{settings.main_db_user}:{settings.main_db_password}@{settings.main_db_host}:{settings.main_db_port}/{settings.main_db_name}"
    sync_engine = create_engine(
        sync_db_url,
        pool_size=5,
        max_overflow=10,
        pool_pre_ping=True,
        echo=False
    )
    
    # Create temporary store filter table
    temp_store_table = f"filter_stores_{int(time.time())}_{id(store_codes)}"
    logger.info(f"🔧 Creating store filter table: {temp_store_table}...")
    
    try:
        # Create temp table and insert stores
        logger.info(f"🔧 Creating temp table {temp_store_table} with {len(store_codes)} stores...")
        with sync_engine.begin() as sync_conn:
            sync_conn.execute(sync_text(f"DROP TABLE IF EXISTS {temp_store_table}"))
            sync_conn.execute(sync_text(f"CREATE TABLE {temp_store_table} (store_code VARCHAR(50) PRIMARY KEY, INDEX idx_store_code (store_code))"))
            
            if len(store_codes) > 0:
                batch_size = 1000
                for i in range(0, len(store_codes), batch_size):
                    batch = store_codes[i:i+batch_size]
                    store_values = ", ".join([f"('{str(store).replace(chr(39), chr(39) + chr(39))}')" for store in batch])
                    sync_conn.execute(sync_text(f"INSERT INTO {temp_store_table} VALUES {store_values}"))
        logger.info(f"✅ Temp table {temp_store_table} created successfully")
        
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Generate Summary sheet first (needs to be first sheet)
        logger.info("📊 Generating Summary sheet with formulas...")
        generate_summary_sheet(workbook, start_date_dt, end_date_dt)
        
        # Helper function to get available columns from a table
        def get_available_columns(table_name, desired_columns):
            """Get columns that actually exist in the table"""
            try:
                check_query = f"SELECT COLUMN_NAME FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = '{table_name}'"
                with sync_engine.connect() as conn:
                    available_cols = conn.execute(sync_text(check_query)).fetchall()
                    available_col_names = {row[0] for row in available_cols}
                    # Return only columns that exist in both lists
                    return [col for col in desired_columns if col in available_col_names]
            except Exception as e:
                logger.warning(f"Could not check columns for {table_name}: {e}")
                return desired_columns  # Fallback to all columns
        
        # Create data sheets
        # 1. Zomato POS vs 3PO
        logger.info("📊 Generating 'Zomato POS vs 3PO' sheet...")
        pos_vs_3po_available_cols = get_available_columns("zomato_pos_vs_3po_data", POS_VS_ZOMATO_COLUMNS)
        pos_vs_3po_columns_str = ", ".join([f"z.{col}" for col in pos_vs_3po_available_cols])
        # Check total count first for debugging
        total_count_query = sync_text("SELECT COUNT(*) as cnt FROM zomato_pos_vs_3po_data")
        with sync_engine.connect() as conn:
            total_count_result = conn.execute(total_count_query).fetchone()
            logger.info(f"   📊 Total records in table: {total_count_result.cnt if total_count_result else 0}")
        
        count_query = sync_text(f"SELECT COUNT(*) as cnt FROM zomato_pos_vs_3po_data WHERE order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'")
        with sync_engine.connect() as conn:
            count_result = conn.execute(count_query).fetchone()
            logger.info(f"   📊 Records in date range ({start_date_dt} to {end_date_dt}): {count_result.cnt if count_result else 0}")
        
        # Build query - include records with NULL store_name OR matching store codes
        if len(store_codes) == 0:
            pos_vs_3po_query = f"""
                SELECT {pos_vs_3po_columns_str}
                FROM zomato_pos_vs_3po_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                ORDER BY z.order_date ASC
            """
        else:
            # Include records with NULL store_name OR records matching store codes
            store_codes_str = "', '".join(store_codes)
            pos_vs_3po_query = f"""
                SELECT {pos_vs_3po_columns_str}
                FROM zomato_pos_vs_3po_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                AND (z.store_name IS NULL OR z.store_name IN ('{store_codes_str}'))
                ORDER BY z.order_date ASC
            """
        logger.info(f"   🔍 Query: {pos_vs_3po_query[:300]}...")
        try:
            with sync_engine.connect() as conn:
                pos_vs_3po_data = conn.execute(sync_text(pos_vs_3po_query)).fetchall()
                pos_vs_3po_dicts = [dict(row._mapping) for row in pos_vs_3po_data] if pos_vs_3po_data else []
            # Use available columns for sheet creation
            create_data_sheet(workbook, "Zomato POS vs 3PO", pos_vs_3po_available_cols, pos_vs_3po_dicts, TEXT_COLUMNS_POS)
            logger.info(f"   ✅ {len(pos_vs_3po_dicts)} rows")
        except Exception as e:
            logger.warning(f"   ⚠️ Error creating Zomato POS vs 3PO sheet: {e}", exc_info=True)
            create_data_sheet(workbook, "Zomato POS vs 3PO", pos_vs_3po_available_cols, [], TEXT_COLUMNS_POS)
        
        # 2. Zomato 3PO vs POS
        logger.info("📊 Generating 'Zomato 3PO vs POS' sheet...")
        zomato_vs_pos_available_cols = get_available_columns("zomato_3po_vs_pos_data", ZOMATO_VS_POS_COLUMNS)
        zomato_vs_pos_columns_str = ", ".join([f"z.{col}" for col in zomato_vs_pos_available_cols])
        total_count_query = sync_text("SELECT COUNT(*) as cnt FROM zomato_3po_vs_pos_data")
        with sync_engine.connect() as conn:
            total_count_result = conn.execute(total_count_query).fetchone()
            logger.info(f"   📊 Total records in table: {total_count_result.cnt if total_count_result else 0}")
        
        count_query = sync_text(f"SELECT COUNT(*) as cnt FROM zomato_3po_vs_pos_data WHERE order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'")
        with sync_engine.connect() as conn:
            count_result = conn.execute(count_query).fetchone()
            logger.info(f"   📊 Records in date range: {count_result.cnt if count_result else 0}")
        
        if len(store_codes) == 0:
            zomato_vs_pos_query = f"""
                SELECT {zomato_vs_pos_columns_str}
                FROM zomato_3po_vs_pos_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                ORDER BY z.order_date ASC
            """
        else:
            store_codes_str = "', '".join(store_codes)
            zomato_vs_pos_query = f"""
                SELECT {zomato_vs_pos_columns_str}
                FROM zomato_3po_vs_pos_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                AND (z.store_name IS NULL OR z.store_name IN ('{store_codes_str}'))
                ORDER BY z.order_date ASC
            """
        try:
            with sync_engine.connect() as conn:
                zomato_vs_pos_data = conn.execute(sync_text(zomato_vs_pos_query)).fetchall()
                zomato_vs_pos_dicts = [dict(row._mapping) for row in zomato_vs_pos_data] if zomato_vs_pos_data else []
            create_data_sheet(workbook, "Zomato 3PO vs POS", zomato_vs_pos_available_cols, zomato_vs_pos_dicts, TEXT_COLUMNS_ZOMATO)
            logger.info(f"   ✅ {len(zomato_vs_pos_dicts)} rows")
        except Exception as e:
            logger.warning(f"   ⚠️ Error creating Zomato 3PO vs POS sheet: {e}", exc_info=True)
            create_data_sheet(workbook, "Zomato 3PO vs POS", zomato_vs_pos_available_cols, [], TEXT_COLUMNS_ZOMATO)
        
        # 3. Zomato 3PO vs POS Refund
        logger.info("📊 Generating 'Zomato 3PO vs POS Refund' sheet...")
        refund_available_cols = get_available_columns("zomato_3po_vs_pos_refund_data", ZOMATO_VS_POS_COLUMNS)
        refund_columns_str = ", ".join([f"z.{col}" for col in refund_available_cols])
        total_count_query = sync_text("SELECT COUNT(*) as cnt FROM zomato_3po_vs_pos_refund_data")
        with sync_engine.connect() as conn:
            total_count_result = conn.execute(total_count_query).fetchone()
            logger.info(f"   📊 Total records in table: {total_count_result.cnt if total_count_result else 0}")
        
        count_query = sync_text(f"SELECT COUNT(*) as cnt FROM zomato_3po_vs_pos_refund_data WHERE order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'")
        with sync_engine.connect() as conn:
            count_result = conn.execute(count_query).fetchone()
            logger.info(f"   📊 Records in date range: {count_result.cnt if count_result else 0}")
        
        if len(store_codes) == 0:
            refund_query = f"""
                SELECT {refund_columns_str}
                FROM zomato_3po_vs_pos_refund_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                ORDER BY z.order_date ASC
            """
        else:
            store_codes_str = "', '".join(store_codes)
            refund_query = f"""
                SELECT {refund_columns_str}
                FROM zomato_3po_vs_pos_refund_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                AND (z.store_name IS NULL OR z.store_name IN ('{store_codes_str}'))
                ORDER BY z.order_date ASC
            """
        try:
            with sync_engine.connect() as conn:
                refund_data = conn.execute(sync_text(refund_query)).fetchall()
                refund_dicts = [dict(row._mapping) for row in refund_data] if refund_data else []
            create_data_sheet(workbook, "Zomato 3PO vs POS Refund", refund_available_cols, refund_dicts, TEXT_COLUMNS_ZOMATO)
            logger.info(f"   ✅ {len(refund_dicts)} rows")
        except Exception as e:
            logger.warning(f"   ⚠️ Error creating Zomato 3PO vs POS Refund sheet: {e}", exc_info=True)
            create_data_sheet(workbook, "Zomato 3PO vs POS Refund", refund_available_cols, [], TEXT_COLUMNS_ZOMATO)
        
        # 4. Order not found in POS
        logger.info("📊 Generating 'Order not found in POS' sheet...")
        not_in_pos_available_cols = get_available_columns("orders_not_in_pos_data", ORDERS_NOT_IN_POS_COLUMNS)
        not_in_pos_columns_str = ", ".join([f"z.{col}" for col in not_in_pos_available_cols])
        total_count_query = sync_text("SELECT COUNT(*) as cnt FROM orders_not_in_pos_data")
        with sync_engine.connect() as conn:
            total_count_result = conn.execute(total_count_query).fetchone()
            logger.info(f"   📊 Total records in table: {total_count_result.cnt if total_count_result else 0}")
        
        count_query = sync_text(f"SELECT COUNT(*) as cnt FROM orders_not_in_pos_data WHERE order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'")
        with sync_engine.connect() as conn:
            count_result = conn.execute(count_query).fetchone()
            logger.info(f"   📊 Records in date range: {count_result.cnt if count_result else 0}")
        
        if len(store_codes) == 0:
            not_in_pos_query = f"""
                SELECT {not_in_pos_columns_str}
                FROM orders_not_in_pos_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                ORDER BY z.order_date ASC
            """
        else:
            store_codes_str = "', '".join(store_codes)
            not_in_pos_query = f"""
                SELECT {not_in_pos_columns_str}
                FROM orders_not_in_pos_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                AND (z.store_name IS NULL OR z.store_name IN ('{store_codes_str}'))
                ORDER BY z.order_date ASC
            """
        try:
            with sync_engine.connect() as conn:
                not_in_pos_data = conn.execute(sync_text(not_in_pos_query)).fetchall()
                not_in_pos_dicts = [dict(row._mapping) for row in not_in_pos_data] if not_in_pos_data else []
            create_data_sheet(workbook, "Order not found in POS", not_in_pos_available_cols, not_in_pos_dicts, TEXT_COLUMNS_ZOMATO)
            logger.info(f"   ✅ {len(not_in_pos_dicts)} rows")
        except Exception as e:
            logger.warning(f"   ⚠️ Error creating Order not found in POS sheet: {e}", exc_info=True)
            create_data_sheet(workbook, "Order not found in POS", not_in_pos_available_cols, [], TEXT_COLUMNS_ZOMATO)
        
        # 5. Order not found in 3PO
        logger.info("📊 Generating 'Order not found in 3PO' sheet...")
        not_in_3po_available_cols = get_available_columns("orders_not_in_3po_data", ORDERS_NOT_IN_3PO_COLUMNS)
        not_in_3po_columns_str = ", ".join([f"z.{col}" for col in not_in_3po_available_cols])
        total_count_query = sync_text("SELECT COUNT(*) as cnt FROM orders_not_in_3po_data")
        with sync_engine.connect() as conn:
            total_count_result = conn.execute(total_count_query).fetchone()
            logger.info(f"   📊 Total records in table: {total_count_result.cnt if total_count_result else 0}")
        
        count_query = sync_text(f"SELECT COUNT(*) as cnt FROM orders_not_in_3po_data WHERE order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'")
        with sync_engine.connect() as conn:
            count_result = conn.execute(count_query).fetchone()
            logger.info(f"   📊 Records in date range: {count_result.cnt if count_result else 0}")
        
        if len(store_codes) == 0:
            not_in_3po_query = f"""
                SELECT {not_in_3po_columns_str}
                FROM orders_not_in_3po_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                ORDER BY z.order_date ASC
            """
        else:
            store_codes_str = "', '".join(store_codes)
            not_in_3po_query = f"""
                SELECT {not_in_3po_columns_str}
                FROM orders_not_in_3po_data z
                WHERE z.order_date BETWEEN '{start_date_dt}' AND '{end_date_dt}'
                AND (z.store_name IS NULL OR z.store_name IN ('{store_codes_str}'))
                ORDER BY z.order_date ASC
            """
        try:
            with sync_engine.connect() as conn:
                not_in_3po_data = conn.execute(sync_text(not_in_3po_query)).fetchall()
                not_in_3po_dicts = [dict(row._mapping) for row in not_in_3po_data] if not_in_3po_data else []
            create_data_sheet(workbook, "Order not found in 3PO", not_in_3po_available_cols, not_in_3po_dicts, TEXT_COLUMNS_POS)
            logger.info(f"   ✅ {len(not_in_3po_dicts)} rows")
        except Exception as e:
            logger.warning(f"   ⚠️ Error creating Order not found in 3PO sheet: {e}", exc_info=True)
            create_data_sheet(workbook, "Order not found in 3PO", not_in_3po_available_cols, [], TEXT_COLUMNS_POS)
        
        # Save workbook
        logger.info(f"💾 Saving workbook to {filepath}...")
        workbook.save(filepath)
        logger.info(f"✅ Workbook saved successfully")
        
    except Exception as e:
        logger.error(f"Error generating summary sheet: {e}", exc_info=True)
        raise
    finally:
        # Cleanup temp table
        try:
            with sync_engine.begin() as sync_conn:
                sync_conn.execute(sync_text(f"DROP TABLE IF EXISTS {temp_store_table}"))
        except Exception as cleanup_error:
            logger.warning(f"Failed to cleanup temp table {temp_store_table}: {cleanup_error}")
